"""Step 2 + Step 4: De-duplicate, apply exclusions, score each record."""
from __future__ import annotations
import re
from collections import Counter
from typing import Iterable
import openpyxl


# ---------- helpers ----------

def cleanstr(s) -> str:
    if s is None:
        return ''
    return str(s).strip()


def mark_in_scope(mt: str) -> bool:
    """Trademark mark-text rule: exact STEALTH-like word OR 'STEM ' + descriptor.

    Note: this is implemented generically based on the supplied root word.
    See `mark_in_scope_for_root()` for the configurable variant.
    """
    return mark_in_scope_for_root(mt, root='STEALTH')


def mark_in_scope_for_root(mark_text: str, root: str) -> bool:
    if not mark_text:
        return False
    mt = mark_text.upper().strip()
    root_u = root.upper().strip()
    if mt == root_u:
        return True
    if mt.startswith(root_u + ' '):
        return True
    if mt.startswith(root_u + '-'):
        return True
    return False


# ---------- Type-aware match helpers (added 10 Jun 2026) --------------------
#
# Operator-driven match logic for word/domain searches. Each search row in
# the form carries a `type` (Exact Match / Starts With / Contains / Similar
# To) and a `phrase`. The four types behave as follows when filtering
# scrape results into the report:
#
#   Exact Match  — mark text equals the phrase (case-insensitive)
#   Starts With  — mark text equals OR starts with `phrase + space|dash`
#   Contains     — the phrase appears anywhere inside the mark text
#   Similar To   — fuzzy match using difflib (ratio >= 0.78 against the
#                  phrase OR against any whitespace-bounded sub-token of
#                  the mark). Threshold picked so 'BAMBOO CONNECT' matches
#                  'BAMBOOCONNECT' and minor typos but not unrelated words.
#
# The two helpers below return True/False; callers (process_trademarks,
# process_companies, process_domains) iterate over the operator's list of
# search criteria and keep records that match ANY row. If the operator
# supplies no rows we fall back to the legacy `mark_in_scope_for_root`
# behaviour so older code paths keep working.

def _fuzzy_ratio(a: str, b: str) -> float:
    """Cheap similarity score using difflib (stdlib, no extra deps)."""
    import difflib
    return difflib.SequenceMatcher(None, a, b).ratio()


def mark_matches_search_row(mark_text: str, search_type: str, phrase: str) -> bool:
    """Single-row match check: does `mark_text` match this (type, phrase)?"""
    if not mark_text or not phrase:
        return False
    mt = (mark_text or '').upper().strip()
    p = (phrase or '').upper().strip()
    if not p:
        return False
    stype = (search_type or '').lower().strip()
    if stype == 'exact match':
        return mt == p
    if stype == 'starts with':
        return mt == p or mt.startswith(p + ' ') or mt.startswith(p + '-')
    if stype == 'contains':
        return p in mt
    if stype == 'similar to':
        # Whole-string ratio
        if _fuzzy_ratio(mt, p) >= 0.78:
            return True
        # Token-level: try ratio against each space-delimited token in the mark
        for tok in mt.split():
            if _fuzzy_ratio(tok, p) >= 0.85:
                return True
        # Token-overlap (10 Jun 2026): "Similar To 'A B C'" also matches if
        # the mark contains any non-trivial (>=4 char) whole-word token from
        # the phrase. This restores compatibility with scraper output where
        # the scraper pre-filtered using individual phrase words: marks like
        # 'Properly Services' should match a search for 'Nicholson Flooring
        # Services' on the SERVICES token. Without this branch, whole-phrase
        # fuzzy matching alone dropped every record. Tokens <4 chars are
        # ignored to avoid trivial overlaps (e.g. 'AND', 'OR', 'AIR' on a
        # three-letter prefix).
        mark_tokens = set(re.findall(r'\w+', mt))
        for ptok in re.findall(r'\w+', p):
            if len(ptok) >= 4 and ptok in mark_tokens:
                return True
        return False
    # Unknown type — fall back to Starts With behaviour (was the old default)
    return mt == p or mt.startswith(p + ' ')


def mark_matches_any(mark_text: str, word_searches: list[dict] | None) -> bool:
    """True if the mark text matches any row in the operator's word_searches.

    Used by `process_trademarks` and `process_companies` to decide which
    scraped records to keep in the report. Empty / None search list returns
    False so callers can fall back to the root-word behaviour.
    """
    if not word_searches:
        return False
    for ws in word_searches:
        if mark_matches_search_row(mark_text, ws.get('type', ''),
                                    ws.get('phrase', '')):
            return True
    return False


def domain_matches_any(domain_text: str, domain_searches: list[dict] | None) -> bool:
    """True if any of the URLs/labels in a domain record matches any row in
    the operator's domain_searches. Re-uses `mark_matches_search_row` because
    domain types are the same four (Exact / Starts With / Contains / Similar
    To) — they just operate on URL strings rather than mark text."""
    if not domain_searches or not domain_text:
        return False
    # Strip URL scheme/www so a "Contains acme.com" rule matches
    # https://www.acme.com/foo. Operators care about the host+path, not the
    # scheme.
    text = str(domain_text)
    for pfx in ('https://', 'http://'):
        if text.lower().startswith(pfx):
            text = text[len(pfx):]
            break
    if text.lower().startswith('www.'):
        text = text[4:]
    for ds in domain_searches:
        if mark_matches_search_row(text, ds.get('type', ''),
                                    ds.get('phrase', '')):
            return True
    return False


def touches_classes(cls_str: str, target: Iterable[int]) -> bool:
    if not cls_str:
        return False
    parts = re.split(r'[,\s]+', str(cls_str))
    nums = [int(p) for p in parts if p.strip().isdigit()]
    return any(n in nums for n in target)


def has_sic(sic_str: str, target_sic: str) -> bool:
    if not sic_str:
        return False
    return target_sic in str(sic_str)


# Legal-entity suffixes stripped before client-name vs owner-name comparison.
# Captures most common UK/US/EU corporate forms. The list is uppercase so
# the caller normalises both sides before checking.
_ENTITY_SUFFIXES = {
    'LIMITED', 'LTD', 'PLC',
    'LLP', 'LLC', 'INC', 'CORP', 'CORPORATION',
    'GMBH', 'AG', 'SA', 'SAS', 'BV', 'NV', 'AB', 'OY',
    'CO', 'COMPANY', 'GROUP', 'HOLDINGS', 'INTERNATIONAL',
}


def _strip_entity_suffixes(name: str) -> str:
    """Strip trailing corporate/legal-entity suffixes from a company name.

    'Friars Airfield Solutions Limited' -> 'FRIARS AIRFIELD SOLUTIONS'
    'Acme Corp. Inc.'                    -> 'ACME'

    Used as a normalisation step before comparing the cited mark's owner
    to the client name — the suffix tail varies (Ltd vs Limited vs plc)
    while the brand-name body is what actually identifies the entity.
    """
    if not name:
        return ''
    tokens = str(name).upper().replace('.', ' ').replace(',', ' ').split()
    while tokens and tokens[-1] in _ENTITY_SUFFIXES:
        tokens.pop()
    return ' '.join(tokens)


def is_likely_client_owner(client_name: str, owner: str) -> bool:
    """Return True when the cited mark's owner name plausibly identifies
    the client themselves (rather than a third party).

    Detection rule (BR-IMG-003, 10 Jun 2026):
      1. Normalise both sides — upper-case + strip legal entity suffixes
         (Limited/Ltd/plc/LLP/LLC/Inc/Corp/Holdings/Group/Co/etc).
      2. Exact match after normalisation -> True.
      3. Fuzzy Jaro-style ratio >= 0.85 over the normalised strings -> True.

    A simple, intentionally narrow detector. The risk of over-flagging
    (a real third-party threat hidden behind a 'Client Likely' label) is
    worse than under-flagging (operator just sees a normal risk row).

    Edge case — 'Bamboo Connect' vs 'Bamboo Connections' fuzzy-matches at
    ≈ 0.88 because the strings share a long prefix; the operator can
    spot the difference at a glance and adjust. Documented trade-off:
    we accept this false positive in exchange for catching real cases
    where the client's own mark name varies slightly from the form input
    (typos, abbreviations, regional suffixes).
    """
    if not client_name or not owner:
        return False
    cn = _strip_entity_suffixes(client_name)
    on = _strip_entity_suffixes(owner)
    if not cn or not on:
        return False
    if cn == on:
        return True
    if _fuzzy_ratio(cn, on) >= 0.85:
        return True
    return False


def is_figurative_type(mark_type: str) -> bool:
    """True when a trademark's Type column denotes a mark with a visual
    element — Figurative, Combined (word + image), Stylised — and therefore
    can pose a Logo-axis threat. Pure 'Word' marks return False.

    Used in process_trademarks to gate Logo-axis row emission in Combined
    reports: a mark with no figurative element doesn't compete on the logo
    side, even if it shares a class with the client's mark.
    """
    if not mark_type:
        return False
    t = str(mark_type).lower()
    return ('figurative' in t) or ('combined' in t) or ('stylized' in t) or ('stylised' in t)


# ---------- BR-011 (11 Jun 2026): Monitoring Report support -----------------

def _parse_search_date(v) -> str:
    """Normalise a Search Date cell value to 'YYYY-MM-DD' or '' on fail.

    Accepts python datetime/date objects (what openpyxl returns for true
    date cells), ISO strings, and the common 'YYYY-MM-DD HH:MM:SS' shape.
    """
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%Y-%m-%d')
        except Exception:
            pass
    s = str(v).strip()
    if not s:
        return ''
    # Strip time portion if present
    return s.split('T')[0].split(' ')[0]


def _find_search_date_col(header) -> int | None:
    """Scan a sheet's header row for a 'Search Date' column.

    Returns the 0-based column index, or None if no such column exists.
    Case-insensitive, tolerant of extra whitespace. Used by
    filter_to_latest_search_date() so monitoring reports can drop rows
    from earlier scrape periods.
    """
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell and 'search date' in str(cell).lower():
            return i
    return None


def filter_to_latest_search_date(rows: list) -> tuple[list, str]:
    """Drop rows whose Search Date isn't the latest in the sheet.

    `rows` is the [header, *data] list of tuples that read_sheets() returns.
    Returns `(filtered_rows, latest_date_iso)`.

    If the sheet has no Search Date column, or every row's date is empty,
    the input is returned unchanged with `latest_date_iso=''`. This makes
    the function safe to apply to every sheet — sheets without dates pass
    through untouched.
    """
    if not rows or len(rows) < 2:
        return rows, ''
    header = rows[0]
    col_idx = _find_search_date_col(header)
    if col_idx is None:
        return rows, ''
    data = rows[1:]
    dates = []
    for r in data:
        if r and len(r) > col_idx:
            d = _parse_search_date(r[col_idx])
            if d:
                dates.append(d)
    if not dates:
        return rows, ''
    latest = max(dates)
    kept = [
        r for r in data
        if r and len(r) > col_idx
        and _parse_search_date(r[col_idx]) == latest
    ]
    return [header] + kept, latest


def extract_order_metadata(xlsx_path: str) -> dict:
    """Read the Order Form sheet and return a dict of order-level metadata
    that can be used to pre-populate the audit UI.

    The parser is label-driven (matches against column A text, case-insensitive,
    whitespace-stripped) rather than row-driven, so changes to the template
    layout don't break it as long as the labels stay the same.

    Returns a dict with whatever it finds, falling back to '' for missing
    fields. Callers should treat the returned values as defaults and let the
    UI surface them for the operator to confirm or override.
    """
    import openpyxl
    out = {
        'client_name': '',
        'deal_id': '',
        'word_or_image': '',
        'sic': '',
        'nature': '',
        'countries': '',
        'search_platforms': '',
        # Back-compat single-value fields (populated from the first matching
        # word_searches / domain_searches entry of that type, so existing UI
        # callers keep working).
        'exact_match': '',
        'starts_with': '',
        'domain_exact': '',
        # New structured fields (added 03 June 2026 for the multi-phrase Order
        # Form). Each entry is {'type': str, 'phrase': str, 'remarks': str}.
        # `type` is the value in column A (e.g. 'Exact Match', 'Starts With',
        # 'Contains', 'Similar To'); `phrase` is the value in column B;
        # `remarks` is the value in column C if present.
        'word_searches': [],
        'domain_searches': [],
        # Image fields (Order Form template rows 30 and 31). image_1 and
        # image_2 are TEXT labels from column B — typically the literal
        # 'NO SEARCH', a filename, the sentinel '<embedded image>' written
        # when an actual image is anchored to the cell, or empty. The raw
        # image bytes (when present) live in image_1_bytes / image_2_bytes
        # so downstream code can render or process them.
        #
        # Vienna classifications block (template update 10 Jun 2026):
        #   D31:D40 — up to 10 Vienna codes in IPO 'NN.NN.NN' format
        #   E31:E40 — matching descriptions
        # `vienna_classifications` is the structured list [{code, description}, ...].
        # `vienna_classes` is the joined codes string kept for back-compat.
        # Legacy spreadsheets that stored Vienna in C30/C31 are still
        # supported via the fallback path below.
        'image_1': '',
        'image_1_bytes': None,    # bytes if an embedded image was anchored to B30
        'image_1_format': '',     # 'jpeg', 'png', etc. — empty if no embedded image
        'image_2': '',
        'image_2_bytes': None,
        'image_2_format': '',
        'vienna_classifications': [],   # NEW structured: [{'code': str, 'description': str}, ...]
        'vienna_classes': '',           # back-compat joined string of codes
        'classes_csv': '',
        'search_date': '',
        # Audit Operator Details block (added to Order Form template rows 58\u201364)
        'brand_reference': '',
        'report_reference': '',
        'client_first': '',
        'client_last': '',
        'client_email': '',
        'account_manager': '',
        'prepared_by': '',
        # BR-011 (11 Jun 2026) — type of deliverable. 'audit' = initial
        # clearance / search-and-audit report. 'monitoring' = scheduled
        # monitoring or representation report (ongoing surveillance).
        # Set from Order Form A2 (the value of the document-type row,
        # below the A1 'Search/Audit or Monitoring & Representation'
        # heading). Default 'audit' is the safe assumption for legacy
        # spreadsheets that don't carry this label.
        'document_type': 'audit',
    }
    try:
        # NB: read_only=True would drop _images, which we need to detect
        # embedded image marks anchored to B30 / B31. The Order Form is small
        # so the regular load is fast enough.
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception:
        return out
    if 'Order Form' not in wb.sheetnames:
        wb.close()
        return out
    ws = wb['Order Form']

    # BR-011: detect monitoring report via Order Form A2 cell. The exact
    # wording in the template is "Monitoring or Representation Report" for
    # monitoring; an audit/search report will have something like
    # "Search/Audit Report" or similar. We match on the substring
    # "monitoring" case-insensitively so cosmetic wording changes don't
    # break detection.
    a2_value = ws.cell(row=2, column=1).value
    if a2_value and 'monitoring' in str(a2_value).lower():
        out['document_type'] = 'monitoring'

    # Map of expected labels (lower-case, stripped) -> key in our out dict.
    # Labels are matched as-is from column A; column B is the value.
    label_map = {
        'client name': 'client_name',
        'deal id': 'deal_id',
        'word or image': 'word_or_image',
        'sic': 'sic',
        'nature of business': 'nature',
        'designated countries': 'countries',
        'search platforms': 'search_platforms',
        # Audit Operator Details block (added to Order Form template rows 58\u201364)
        'brand reference': 'brand_reference',
        'report reference': 'report_reference',
        'client first name': 'client_first',
        'client last name': 'client_last',
        # Hopper-style template (Jun 2026) renamed these labels:
        'contact first name': 'client_first',
        'contact last name': 'client_last',
        'client email address': 'client_email',
        'account manager': 'account_manager',
        'report prepared by': 'prepared_by',
    }

    # Column-B values that are template instructions / CRM source-field mappings
    # rather than real client data \u2014 treat as empty when seen, so the UI shows
    # blank required fields instead of pre-filling literal "Contact First Name".
    placeholder_values = {
        'add field',
        'add single field',
        'add single field - usually word mark text',
        'contact first name',
        'contact last name',
        'contact email',
        'contact email address',
        'deal owner',
        'deal name',
        'pipeline name',
        '{{pipeline name}} report',
        'headshot',  # R42 placeholder for an image cell in the legacy template
        'completed by tmh',
        'completed by suntec of tmh',
        # Hopper-template placeholders (Jun 2026) — these are template-stencil
        # strings the operator should overwrite. When seen, treat as empty so
        # downstream code can render "(not specified)" rather than leaking the
        # placeholder into the report.
        'type',                                       # row 14 B placeholder
        'sic code',                                   # row 15 B placeholder
        'additional sic code, nature of business',    # row 16 B placeholder
        'tm search platforms',                        # row 17 B placeholder
        'other search platforms',                     # row 18 B placeholder
        'applicant account name',                     # Order Form (2) wireframe
        'crm deal id',                                # Order Form (2) wireframe
        'image mark information/image jpeg',          # Order Form (2) wireframe
        'added manually post search',                 # operator-typed "fill in later" marker
    }
    def _real_value(b) -> str:
        if b is None:
            return ''
        s = str(b).strip()
        if s.lower() in placeholder_values:
            return ''
        if s.lower().startswith('add '):  # 'Add Field', 'Add Single Field...'
            return ''
        return s
    # Search-criteria block uses 'Exact Match' as a label, which appears
    # twice (once for Text, once for Domain) so we need to track which
    # context we're in via the preceding header row.
    text_ctx = False
    domain_ctx = False

    # Recognised search-criterion types in column A of the words/domains
    # blocks. The map is many → one so different template generations can
    # all funnel into the canonical four labels. Anything outside this set
    # (e.g. 'Contains2' from numbered template slots) is treated as 'not a
    # recognised search row' and skipped.
    SEARCH_TYPE_ALIASES = {
        'exact match':  'Exact Match',
        'starts with':  'Starts With',
        'start with':   'Starts With',     # Friars template typo / variant
        'contains':     'Contains',
        'similar to':   'Similar To',
        'similar match': 'Similar To',     # Friars template variant
    }

    # Class list rows start with a class number like '11 - Heating Components'.
    # Capture flips ON when the "G&S Classes" header is seen (templates put
    # this block at varying rows — R26 in Friars, R33 in Woodcross).
    import re
    class_nums: list[int] = []
    g_s_classes_seen = False

    for r in range(1, min(ws.max_row + 1, 100)):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a is None:
            continue
        a_str = str(a).strip()
        # Normalize: drop trailing required-field asterisk and surrounding spaces
        # so 'Client first name *' matches 'client first name'.
        a_key = a_str.rstrip('*').strip().lower()

        # Direct label-value matches.
        # BR-012 (12 Jun 2026): "first non-empty wins" — the Hopper-style
        # template carries duplicate label blocks (a populated block at
        # rows 5-18 AND a legacy wireframe block at rows 58-71 with
        # placeholder text). Without this guard the legacy block would
        # overwrite the real values with placeholders. Once a field has
        # been set to a real value, later rows with the same label are
        # silently skipped.
        if a_key in label_map:
            field = label_map[a_key]
            new_val = _real_value(b)
            existing = out.get(field)
            if not existing:                  # field still empty → set it
                out[field] = new_val
            # If the field already has a value, only overwrite when the
            # new value is non-empty too — but never replace a real value
            # with a placeholder/empty.
            elif new_val and not existing.strip():
                out[field] = new_val
            continue

        # Context switching for the dual 'Exact Match' labels. Multiple
        # template generations are supported:
        #   "Search Criteria (Text)"     (Woodcross / older)
        #   "Word Search Criteria"       (Friars / newer)
        # Same for domain headers.
        is_text_header = (
            ('search criteria' in a_key and 'text' in a_key) or
            'word search criteria' in a_key or
            ('word search' in a_key and 'criteria' not in a_key)
        )
        is_domain_header = (
            ('search criteria' in a_key and 'domain' in a_key) or
            'domain search criteria' in a_key
        )
        if is_text_header:
            text_ctx, domain_ctx = True, False
            continue
        if is_domain_header:
            text_ctx, domain_ctx = False, True
            continue
        if 'image search' in a_key or 'g&s classes' in a_key:
            text_ctx = domain_ctx = False
            # Once we see the G&S Classes header, mark subsequent rows as
            # eligible for class-number capture. Different templates put
            # this block at different rows (R26 in Friars, R33 in Woodcross),
            # so an absolute-row threshold no longer works.
            if 'g&s classes' in a_key:
                g_s_classes_seen = True
            # fall through so we can still match 'image mark 1' / 'image mark 2'
            # against the explicit handlers below.

        # Words block: capture every (type, phrase) pair, not just Exact /
        # Starts With. Phrase must be non-empty; remarks (column C) optional.
        # SEARCH_TYPE_ALIASES canonicalises template variants like "Start
        # With" → "Starts With" and "Similar Match" → "Similar To".
        canonical_type = SEARCH_TYPE_ALIASES.get(a_key)
        if text_ctx and canonical_type:
            phrase = (str(b).strip() if b is not None else '')
            if phrase:
                remarks = (str(c).strip() if c is not None else '')
                out['word_searches'].append({
                    'type': canonical_type,
                    'phrase': phrase,
                    'remarks': remarks,
                })
                # Back-compat: populate single-value fields from the first
                # matching entry of each type so the existing UI doesn't break.
                if canonical_type == 'Exact Match' and not out['exact_match']:
                    out['exact_match'] = phrase
                elif canonical_type == 'Starts With' and not out['starts_with']:
                    out['starts_with'] = phrase
            continue

        # Domains block: same shape as words.
        if domain_ctx and canonical_type:
            phrase = (str(b).strip() if b is not None else '')
            if phrase:
                remarks = (str(c).strip() if c is not None else '')
                out['domain_searches'].append({
                    'type': canonical_type,
                    'phrase': phrase,
                    'remarks': remarks,
                })
                if canonical_type == 'Exact Match' and not out['domain_exact']:
                    out['domain_exact'] = phrase
            continue

        # Image rows: 'Image Mark 1' (R30) and 'Image Mark 2' (R31). The
        # value cell B may be a filename, the literal 'NO SEARCH' sentinel,
        # or empty. Column C carries the shared Vienna classification codes.
        if a_key in ('image mark 1', 'image mark 2'):
            value = (str(b).strip() if b is not None else '')
            # 'NO SEARCH' is a legitimate signal that this image slot is
            # intentionally unused; preserve it verbatim rather than blanking
            # it, so downstream code can distinguish "no image search wanted"
            # from "image search wanted but image filename missing".
            slot = 'image_1' if a_key == 'image mark 1' else 'image_2'
            out[slot] = value
            vienna = (str(c).strip() if c is not None else '')
            if vienna and not out['vienna_classes']:
                out['vienna_classes'] = vienna
            continue

        # Class row: 'NN - Description'. Only capture after the G&S Classes
        # header has been seen — different templates put this block at very
        # different rows so an absolute-row threshold is unreliable.
        m = re.match(r'^(\d{1,2})\b', a_str)
        if m and g_s_classes_seen:
            class_nums.append(int(m.group(1)))
            continue

        # Date row \u2014 only accept actual date-typed values. The current
        # Order Form template uses placeholder strings ('Date Report Produced',
        # 'Completed by TMH') in this block; if the value isn't a real date,
        # we leave it blank and the UI defaults to today.
        if 'date of most recent search' in a_key:
            from datetime import date, datetime
            candidates = [b, ws.cell(row=r+1, column=1).value, ws.cell(row=r+1, column=2).value]
            for cand in candidates:
                if isinstance(cand, (date, datetime)):
                    out['search_date'] = cand.strftime('%d %B %Y')
                    break
            continue

    if class_nums:
        out['classes_csv'] = ', '.join(str(n) for n in class_nums)

    # Embedded images on the Order Form sheet — operator pastes the actual
    # logo image into the cell next to 'Image Mark 1' / 'Image Mark 2'.
    # BR-012 (12 Jun 2026): switched from hardcoded rows 30/31 to dynamic
    # label discovery, because the Hopper-style template moved this block
    # to rows 37/38. We scan column A for the labels first, then match
    # image anchors to the resolved rows (column B, +/-1 for tall images).
    image_label_rows: dict[str, int] = {}  # 'image_1' or 'image_2' -> 1-based row
    for r in range(1, min(ws.max_row + 1, 100)):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        a_norm = str(a).rstrip('*').strip().lower()
        if a_norm == 'image mark 1' and 'image_1' not in image_label_rows:
            image_label_rows['image_1'] = r
        elif a_norm == 'image mark 2' and 'image_2' not in image_label_rows:
            image_label_rows['image_2'] = r
    # Back-compat: if no labels found, fall back to legacy rows 30/31.
    if not image_label_rows:
        image_label_rows = {'image_1': 30, 'image_2': 31}

    for img in (getattr(ws, '_images', None) or []):
        try:
            anchor = img.anchor
            from_ = getattr(anchor, '_from', None)
            if from_ is None:
                continue
            # openpyxl uses 0-indexed AnchorMarker; col 1 == B.
            if from_.col != 1:
                continue
            row_1based = from_.row + 1
            # Match the anchor row to whichever 'Image Mark N' label is closest
            # (within ±2 rows — Excel anchors are sometimes off-by-one for tall
            # images that bleed into adjacent cells).
            slot = None
            for candidate_slot, label_row in image_label_rows.items():
                if abs(row_1based - label_row) <= 2:
                    slot = candidate_slot
                    break
            if slot is None:
                continue
            # Pull raw bytes — _data() is the openpyxl Image API
            try:
                blob = img._data()
                data = bytes(blob) if isinstance(blob, (bytes, bytearray)) else None
            except Exception:
                data = None
            if not data:
                continue
            out[f'{slot}_bytes'] = data
            out[f'{slot}_format'] = (getattr(img, 'format', '') or '').lower()
            # Only promote the cell label if the text cell was empty —
            # don't overwrite an explicit operator note or filename.
            if not out[slot]:
                out[slot] = '<embedded image>'
        except Exception:
            # Image walk is best-effort; an unparseable image shouldn't
            # break the rest of the order-form parse.
            continue

    # ---- Vienna classifications block (pattern scan of column D) ----------
    # Different template generations put Vienna data at different rows:
    #   Older templates: D31:D40 + E31:E40
    #   Friars template: D23:D34 (inline with the Image section)
    # Rather than hard-coding rows, scan column D for cells that match the
    # Vienna code pattern (digits separated by 1-2 dots, e.g. "29.1.8" or
    # "26.13.25"). Capture the matching column-E cell as the description.
    # Skip header rows ("Classification", "Description", etc.).
    vienna_code_re = re.compile(r'^\d+\.\d+(\.\d+)?$')
    new_vienna: list[dict] = []
    for r in range(1, min(ws.max_row + 1, 60)):
        code = ws.cell(row=r, column=4).value   # column D
        if code is None:
            continue
        code_str = str(code).strip()
        # Skip blank cells and header text
        if not code_str:
            continue
        if code_str.lower() in (
            'classification', 'classifications',
            'image class.division.subdivision',
            'description', 'descriptions',
        ):
            continue
        # Only accept values that look like Vienna codes
        if not vienna_code_re.match(code_str):
            continue
        desc = ws.cell(row=r, column=5).value   # column E
        desc_str = str(desc).strip() if desc is not None else ''
        new_vienna.append({'code': code_str, 'description': desc_str})
    if new_vienna:
        out['vienna_classifications'] = new_vienna
        # Overwrite the legacy joined string with the new authoritative list.
        out['vienna_classes'] = ', '.join(v['code'] for v in new_vienna if v.get('code'))
    elif out.get('vienna_classes'):
        # Legacy spreadsheets using only C30/C31. Promote the joined string
        # to a single-entry structured list so downstream code can iterate.
        out['vienna_classifications'] = [{'code': out['vienna_classes'], 'description': ''}]

    wb.close()
    return out


def extract_google_image_cells(xlsx_path: str) -> dict[int, bytes]:
    """Extract Excel 'Picture in Cell' images from the Google sheet, column A.

    Modern Excel (365+) supports inserting images directly INTO a cell as a
    rich value (the so-called "Picture in Cell" feature). Cells using this
    feature look like Excel error cells to most XLSX readers — openpyxl
    reports them as `#VALUE!` — but the image bytes are present in
    `xl/media/imageN.png` and referenced via `xl/richData/richValueRel.xml`.

    Each affected cell carries a `vm="N"` attribute on its `<c>` element
    where N is a 1-based index into richValueRel. richValueRel.xml has
    `<rel r:id="rIdX"/>` entries in order, and `xl/richData/_rels/
    richValueRel.xml.rels` maps each rId to a path under xl/media/.

    Returns {row_number: image_bytes} for every column-A cell on the
    Google sheet that resolves to a rich-value image. Returns {} if the
    workbook has no rich values, the Google sheet is missing, or any
    parse step fails — failure is silent so a malformed file never crashes
    the audit pipeline.

    Why direct zip+XML parsing rather than openpyxl: openpyxl's high-level
    cell API exposes the cached error value but not the `vm` attribute,
    and `ws._images` only contains drawing-anchored images, not rich-value
    images. We have to read the underlying OOXML structures directly.
    """
    out: dict[int, bytes] = {}
    try:
        import zipfile
        import xml.etree.ElementTree as ET
        with zipfile.ZipFile(xlsx_path) as zf:
            # ---- Step 1: ordered list of rIds from richValueRel.xml ----
            try:
                rel_xml = zf.read('xl/richData/richValueRel.xml')
                rels_xml = zf.read('xl/richData/_rels/richValueRel.xml.rels')
            except KeyError:
                return out  # workbook has no rich values
            ns_r = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
            rel_root = ET.fromstring(rel_xml)
            rel_ids: list[str] = []  # index 0 corresponds to vm="1"
            for child in rel_root:
                rid = child.get(f'{ns_r}id')
                if rid:
                    rel_ids.append(rid)

            # ---- Step 2: map rId -> image path inside the zip ----
            rels_root = ET.fromstring(rels_xml)
            id_to_image_path: dict[str, str] = {}
            for rel in rels_root:
                rid = rel.get('Id')
                target = rel.get('Target') or ''
                if not rid or not target:
                    continue
                # Targets are relative to xl/richData/, e.g. "../media/image1.png"
                if target.startswith('../'):
                    full = 'xl/' + target[3:]
                elif target.startswith('/'):
                    full = target.lstrip('/')
                else:
                    full = 'xl/richData/' + target
                id_to_image_path[rid] = full

            # ---- Step 3: locate the Google sheet's XML file via rels ----
            # Use the rId from xl/workbook.xml + xl/_rels/workbook.xml.rels so
            # we don't depend on the sheet's position number matching its
            # filename (Excel sometimes diverges).
            ns_x = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
            wb_xml = ET.fromstring(zf.read('xl/workbook.xml'))
            google_rid = None
            for sheet in wb_xml.findall(f'{ns_x}sheets/{ns_x}sheet'):
                if (sheet.get('name') or '').strip().lower() == 'google':
                    google_rid = sheet.get(f'{ns_r}id')
                    break
            if not google_rid:
                return out
            wb_rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            sheet_path = None
            for rel in wb_rels:
                if rel.get('Id') == google_rid:
                    target = rel.get('Target') or ''
                    # Target like "worksheets/sheet4.xml" — prefix with xl/
                    if target.startswith('/'):
                        sheet_path = target.lstrip('/')
                    else:
                        sheet_path = 'xl/' + target
                    break
            if not sheet_path:
                return out
            try:
                sheet_xml = zf.read(sheet_path)
            except KeyError:
                return out

            # ---- Step 4: walk column-A cells with a vm attribute ----
            sheet_root = ET.fromstring(sheet_xml)
            for row in sheet_root.findall(f'{ns_x}sheetData/{ns_x}row'):
                for cell in row.findall(f'{ns_x}c'):
                    ref = cell.get('r') or ''  # e.g. "A5"
                    if not ref.startswith('A') or len(ref) < 2 or not ref[1:].isdigit():
                        continue
                    row_num = int(ref[1:])
                    vm = cell.get('vm')
                    if not vm:
                        continue
                    try:
                        vm_idx = int(vm) - 1
                    except ValueError:
                        continue
                    if not (0 <= vm_idx < len(rel_ids)):
                        continue
                    img_path = id_to_image_path.get(rel_ids[vm_idx])
                    if not img_path:
                        continue
                    try:
                        blob = zf.read(img_path)
                    except KeyError:
                        continue
                    if blob:
                        out[row_num] = blob
    except Exception:
        # Best-effort extraction — any parse failure simply means no images.
        return out
    return out


def extract_trademark_images(xlsx_path: str) -> dict[int, bytes]:
    """Pull floating images anchored to column E (the 'Image' column) of the
    Trademarks sheet.

    Returns a dict mapping the 1-based row number of the Trademarks sheet to
    the raw image bytes (typically JPEG). Each trademark mark image is
    embedded in the workbook by the scraper as an anchored picture, not as
    text in the cell, so we have to walk the worksheet's _images list rather
    than reading the cell value.

    Returns {} if the Trademarks sheet has no images or can't be opened.
    """
    out: dict[int, bytes] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)  # NB: read_only=True drops _images
    except Exception:
        return out
    if 'Trademarks' not in wb.sheetnames:
        wb.close()
        return out
    ws = wb['Trademarks']
    for img in (getattr(ws, '_images', None) or []):
        # Determine the row this image is anchored to. openpyxl uses a
        # zero-indexed AnchorMarker (.col, .row); we want 1-based row.
        row_num = None
        try:
            anchor = img.anchor
            from_ = getattr(anchor, '_from', None)
            if from_ is not None:
                row_num = from_.row + 1
        except Exception:
            row_num = None
        if row_num is None:
            continue
        # Pull the raw bytes. openpyxl's Image wraps either a PIL image or
        # a file-like / blob handle. The most reliable cross-version path is
        # to read .ref or ._data() depending on the openpyxl version.
        data = None
        try:
            blob = img._data()  # method on Image \u2014 returns bytes
            if isinstance(blob, (bytes, bytearray)):
                data = bytes(blob)
        except Exception:
            try:
                ref = getattr(img, 'ref', None)
                if ref is not None and hasattr(ref, 'read'):
                    data = ref.read()
            except Exception:
                data = None
        if data:
            out[row_num] = data
    wb.close()
    return out


def extract_specific_terms(xlsx_path: str) -> dict[int, str]:
    """Read the Order Form sheet and pull the client's specific Goods & Services
    terms per class.

    The Braudit order form lists classes immediately after the "G&S Classes"
    header in the format:
        Column A: '11 - Heating Components'
        Column B: 'LED light strips; LED underwater lights; ...'

    Different templates place this block at different rows (R34+ in
    Woodcross-style templates, R27+ in Friars-style templates) so we scan
    for the "G&S Classes" header first and start capturing from the row
    after. Parsing stops at the first empty A cell after we've started
    capturing, or when A starts with anything other than a class-number-
    prefixed entry (e.g. 'Date of Most Recent Search').

    Returns a dict mapping class_number -> specific terms text.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if 'Order Form' not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb['Order Form']
    out: dict[int, str] = {}

    # First pass: find the G&S Classes header row
    header_row = None
    for r in range(1, min(ws.max_row + 1, 80)):
        a = ws.cell(row=r, column=1).value
        if a and 'g&s classes' in str(a).strip().lower():
            header_row = r
            break
    if header_row is None:
        wb.close()
        return out

    # Second pass: capture class-number rows starting just after the header
    started = False
    for r in range(header_row + 1, min(ws.max_row + 1, header_row + 30)):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a is None:
            # Empty row \u2014 keep scanning a few more in case the header is
            # immediately followed by a blank, but once we've started
            # capturing, an empty row ends the block.
            if started:
                break
            continue
        a_str = str(a).strip()
        m = re.match(r'^(\d{1,2})\b', a_str)
        if not m:
            if started:
                break  # non-class row encountered after the block
            continue
        cls = int(m.group(1))
        terms = str(b).strip() if b else ''
        out[cls] = terms
        started = True
    wb.close()
    return out


# ---------- main pipeline ----------

def read_sheets(xlsx_path: str) -> dict:
    """Read every supported sheet of the scraped-results workbook into rows lists."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets = {}
    for name in ['Order Form', 'Google', 'Companies', 'Domains', 'Social', 'Trademarks']:
        if name in wb.sheetnames:
            ws = wb[name]
            sheets[name] = list(ws.iter_rows(values_only=True))
    wb.close()
    return sheets


def _mark_similarity_score(mark_u: str, word_searches: list[dict] | None,
                           root: str) -> int:
    """Return the mark-similarity component of the word score.

    BR-013 (19 Jun 2026): replace the single-root "exact or starts-with"
    test with a per-search-row scan. The operator's word_searches list
    drives which rows of cited marks land in the report (handled in the
    filter layer via mark_matches_search_row). This function gives those
    same matches a SCORING reward — previously only Exact-or-StartsWith
    matches against the first word of the first Exact phrase scored
    anything, so "Similar To" and "Contains" hits got into the report but
    were graded as if their mark text was unrelated to the brand. That
    explains why marks like MOMENTUM MORTGAGE and Momentous (1-letter
    variants of MOMENTUS in class 36) all sat at Low Risk despite being
    obvious threats.

    Points awarded (max across all matching rows):
        Exact Match     +4   (mark equals an Exact phrase)
        Starts With     +2   (mark begins with a Starts-With phrase + sep)
        Contains        +2   (Contains phrase appears as a whole word)
        Similar (>=.85) +2   (whole-string OR token-level fuzzy)
        Similar (>=.78) +1   (whole-string OR token-level fuzzy)

    Falls back to the legacy root-based behaviour when no word_searches
    are supplied (preserves back-compat with legacy Step-5 callers).
    """
    if not word_searches:
        # Legacy path: original Exact-or-StartsWith-root behaviour.
        root_u = (root or '').upper().strip()
        if not root_u:
            return 0
        if mark_u == root_u:
            return 4
        if mark_u.startswith(root_u + ' ') or mark_u.startswith(root_u + '-'):
            return 2
        return 0

    best = 0
    for ws in word_searches:
        stype = (ws.get('type', '') or '').lower().strip()
        phrase = (ws.get('phrase', '') or '').upper().strip()
        if not phrase:
            continue
        if stype == 'exact match':
            if mark_u == phrase:
                best = max(best, 4)
        elif stype == 'starts with':
            if (mark_u == phrase or mark_u.startswith(phrase + ' ')
                    or mark_u.startswith(phrase + '-')):
                best = max(best, 2)
        elif stype == 'contains':
            # Whole-word boundary so 'MOMENTUS' doesn't match 'INSTRUMENTUS'.
            if re.search(r'\b' + re.escape(phrase) + r'\b', mark_u):
                best = max(best, 2)
        elif stype == 'similar to':
            whole = _fuzzy_ratio(mark_u, phrase)
            if whole >= 0.85:
                best = max(best, 2)
            elif whole >= 0.78:
                best = max(best, 1)
            # Token-level: a single mark token close to the phrase is the
            # common case (e.g. mark 'MOMENTUM MORTGAGE' vs phrase
            # 'MOMENTUS' — token 'MOMENTUM' fuzzy-matches at ~0.875).
            for tok in mark_u.split():
                tr = _fuzzy_ratio(tok, phrase)
                if tr >= 0.85:
                    best = max(best, 2)
                elif tr >= 0.78:
                    best = max(best, 1)
    return best


def score_trademark(r: tuple, target_classes=(11, 12, 35), root='STEALTH',
                    word_searches: list[dict] | None = None) -> int:
    """Initial-review WORD score on data alone.

    The mark-similarity component is now driven by `word_searches` (the
    operator's per-row Exact / Starts With / Contains / Similar To list)
    when supplied. When `word_searches` is None or empty, falls back to
    the legacy root-word behaviour so older callers keep working.
    """
    status = cleanstr(r[2]).lower()
    mark = cleanstr(r[5]).upper()
    mtype = cleanstr(r[3])
    classes = cleanstr(r[7])
    parts = [int(p) for p in re.split(r'[,\s]+', classes) if p.strip().isdigit()]
    overlap = sum(1 for n in parts if n in target_classes)

    score = 0
    if status == 'registered':
        score += 4
    elif status == 'pending':
        score += 3
    # 'ended' contributes 0

    score += _mark_similarity_score(mark, word_searches, root)

    if mtype.lower() == 'word':
        score += 2
    elif mtype.lower() == 'combined':
        score += 1
    elif 'stylized' in mtype.lower():
        score += 1

    score += overlap
    return score


def score_trademark_image(r: tuple, target_classes=(11, 12, 35),
                          client_vienna: str = '') -> int:
    """Initial-review IMAGE score on data alone.

    Combined reports need a separate image-threat score independent of the
    word axis. Per the 10 Jun 2026 spec, Vienna code overlap is NOT used at
    initial-audit time — cited-record Vienna codes are fetched from WIPO
    only during the forensic layer (see BR-009). Initial audit image risk
    therefore relies on type + class overlap + status.

    Scoring out of ~10 to match the same risk_from_score bands as the word
    axis (the 'mark text similarity' component is replaced by 'figurative
    nature', so pure-word cited marks land near 0 on the image axis):
        Status:   registered +4, pending +3, ended 0
        Type:     figurative +4, combined +3, stylised +2, word 0
        Classes:  +1 per overlapping Nice class (capped via overlap count)

    `r` is the same Trademarks-sheet row tuple consumed by score_trademark.
    `client_vienna` is accepted for forward compatibility — when the forensic
    layer eventually feeds verified cited-record Vienna codes through, a
    Vienna-overlap component will be added here. Today it's a documented
    no-op so initial-audit scoring stays deterministic.
    """
    status = cleanstr(r[2]).lower()
    mtype = cleanstr(r[3]).lower()
    classes = cleanstr(r[7])
    parts = [int(p) for p in re.split(r'[,\s]+', classes) if p.strip().isdigit()]
    overlap = sum(1 for n in parts if n in target_classes)

    score = 0
    if status == 'registered':
        score += 4
    elif status == 'pending':
        score += 3
    # 'ended' contributes 0 (will fall to Negligible regardless)

    # Type component: figurative marks are the primary image threat;
    # combined marks (word+image) are next; word-only marks do NOT threaten
    # an image registration so they score 0 here.
    if 'figurative' in mtype:
        score += 4
    elif 'combined' in mtype:
        score += 3
    elif 'stylized' in mtype or 'stylised' in mtype:
        score += 2
    # word-only: 0

    score += overlap

    # Vienna code overlap: cited record Vienna codes aren't currently
    # captured by the scrape, so this is a graceful no-op for now. If the
    # cited record exposes them in a future pipeline upgrade (e.g. via
    # Signa's mark_feature_type detail or a separate vienna_codes field),
    # we'll add a +2 per overlapping Vienna group up to a cap.
    # The hook is here; the data isn't yet.
    _ = client_vienna  # placeholder for future Vienna scoring

    return score


def risk_from_score(score: int, status: str) -> str:
    if status.lower() == 'ended':
        return 'Negligible'
    if score >= 11:
        return 'High Risk'
    if score >= 8:
        return 'Medium Risk'
    return 'Low Risk'


def report_type_from_meta(order_meta: dict) -> str:
    """Return one of 'word_only' | 'logo' | 'combined' based on order_meta.

    The Order Form template's B7 cell carries the report-type indicator:
        "Word Only"                  -> word_only
        "Logo or Figurative Mark"    -> logo
        "Combined Word and Logo"     -> combined
    Anything else (legacy / unrecognised) defaults to word_only.
    """
    label = (order_meta.get('word_or_image') or '').strip().lower()
    if 'combined' in label:
        return 'combined'
    if 'logo' in label or 'figurative' in label or 'image' in label:
        return 'logo'
    return 'word_only'


def process_trademarks(rows: list, target_classes=(11, 12, 35), root='STEALTH',
                       images: dict[int, bytes] | None = None,
                       word_searches: list[dict] | None = None,
                       report_type: str = 'word_only',
                       client_vienna: str = '',
                       client_image_bytes: bytes | None = None,
                       client_name: str = '') -> list[dict]:
    """Steps 2-4 on the Trademarks sheet.

    `images` is an optional {row_number: jpeg_bytes} map from
    extract_trademark_images(); when supplied each scored record carries the
    matching image bytes (if any) under the 'image_bytes' key so the report
    builder can render the mark image inline.

    `word_searches` is the operator's list of {type, phrase, remarks} rows
    from the audit form. When supplied, marks are filtered with type-aware
    matching (Exact Match / Starts With / Contains / Similar To) against any
    row. When omitted, falls back to the legacy `root`-prefix behaviour so
    older callers keep working.

    `report_type` is one of 'word_only' | 'logo' | 'combined' from
    report_type_from_meta(). Determines whether each cited record produces
    one row or two:
        word_only -> one row per record, risk = word_risk
        logo      -> one row per record, risk = image_risk
        combined  -> two rows when BOTH axes are non-Negligible (Low or
                     higher), each tagged with 'threat_type' = 'Word' or
                     'Logo'. Single row when only one axis is non-Negligible.
                     If both are Negligible, single row labelled by the
                     higher-scoring axis (ties → Word).

    `client_vienna` is the operator's Vienna code string for the client
    image (from order_meta.vienna_classes). Used by image scoring.
    """
    images = images or {}
    # Carry the original spreadsheet row number alongside each data row so we
    # can look up the corresponding image after dedupe / filter / sort.
    # Header is row 1; data starts at row 2.
    data = [(idx + 2, r) for idx, r in enumerate(rows[1:]) if r[0] is not None]

    # Dedupe by (Office, App Number) \u2014 keep the first occurrence (which
    # preserves the earliest image too).
    seen = set()
    deduped = []
    for row_num, r in data:
        key = (cleanstr(r[0]), cleanstr(r[1]))
        if key in seen:
            continue
        seen.add(key)
        deduped.append((row_num, r))

    # Apply exclusion rules. Two things changed here on 10 Jun 2026:
    #
    # 1. word_searches is NO LONGER a record-level filter. The scraper
    #    already did relevance filtering (it brought back marks matching
    #    individual words from the phrase, e.g. "AIRFIELD" / "FRIARS" /
    #    "SOLUTIONS" rows for a "Friars Airfield Solutions" search). My
    #    previous filter required the full phrase to match every mark,
    #    which dropped nearly every record. word_searches still drives
    #    scoring (via root_word in score_trademark) so editing the phrase
    #    in the form still changes which records float to High Risk —
    #    but no record gets excluded from the report.
    # 2. The legacy `mark_in_scope_for_root` fallback only fires now when
    #    word_searches is empty AND a non-default root was passed. This
    #    preserves back-compat with older callers (e.g. legacy Step-5
    #    scripts) without applying it on every modern commission.
    filtered = []
    use_legacy_root_filter = (not word_searches) and root and root != 'STEALTH'
    for row_num, r in deduped:
        mark = cleanstr(r[5])
        classes = cleanstr(r[7])
        if use_legacy_root_filter and not mark_in_scope_for_root(mark, root):
            continue
        if not touches_classes(classes, target_classes):
            continue
        filtered.append((row_num, r))

    # NOTE: BR-IMG-001 (logo OCR for Word-axis recall) was investigated and
    # withdrawn on 10 Jun 2026 — the existing scrapers already extract mark
    # text into the spreadsheet's mark column for all our test templates
    # (Friars, Nicholson, Bamboo), so OCR-on-cited-marks added ~55 sec of
    # audit overhead for 0 measurable Word recoveries. The logo_ocr.py
    # module is kept on disk for future use (client-logo pre-flight check,
    # Google-image-result OCR) but is NOT wired into this scoring path.

    # ---------- BR-IMG-002: Visual similarity (Phase 2, 10 Jun 2026) ------
    # Compute a single batch of visual-similarity decisions for the
    # client logo vs every cited mark logo. Used downstream in the scoring
    # loop to selectively un-cap the Logo-axis risk grade (which is
    # otherwise capped at Medium without any visual signal).
    #
    # Rule (per Jonathan, 10 Jun 2026):
    #   decision='identical' + Live  + class overlap → High Risk
    #   decision='identical' + Live  + no class overlap → Medium Risk
    #   decision='identical' + Ended (Dead) → Medium Risk (upgrade)
    #   decision='similar'  → keep base (Medium-capped) score
    #   decision='weak' or 'unrelated' (CLIP definitively different) → Low Risk
    #   decision='skipped' (model unavailable, image decode failed) → keep base
    #
    # Visual similarity is only computed for Combined / Image reports
    # where the operator supplied a client image. Word-only reports do
    # NOT load the model — saves ~340 MB resident memory and ~30 sec
    # audit time when visual signal isn't needed.
    visual_decisions: dict[int, dict] = {}
    if client_image_bytes and images and report_type in ('logo', 'combined'):
        try:
            from .visual_similarity import score_visual_similarity_batch
            cited_for_visual = {row_num: images[row_num]
                                for row_num, _ in filtered
                                if images.get(row_num)}
            if cited_for_visual:
                visual_decisions = score_visual_similarity_batch(
                    client_image_bytes, cited_for_visual,
                    auto_unload=True,
                )
        except Exception:
            # visual_similarity module unavailable (e.g. open_clip not
            # installed in this environment). Skip the visual layer —
            # audit continues with base Logo-axis scoring + Medium cap.
            visual_decisions = {}

    # Score every record on BOTH axes (word + image). Single-axis reports
    # pick one; combined reports emit one or two rows per record per the
    # duplication rule.
    scored = []
    for row_num, r in filtered:
        status_str = cleanstr(r[2])
        word_score = score_trademark(r, target_classes, root,
                                     word_searches=word_searches)
        image_score = score_trademark_image(r, target_classes, client_vienna)
        word_risk = risk_from_score(word_score, status_str)
        image_risk = risk_from_score(image_score, status_str)

        # ----- BR-IMG-002: cap Logo-axis at Medium without visual signal --
        # Without CLIP, the base score happily hits High Risk on any
        # Registered Figurative mark with 3-class overlap — a known false
        # positive. Cap to Medium here; the visual layer below may
        # un-cap back to High when CLIP confirms genuine similarity.
        if image_risk == 'High Risk':
            image_risk = 'Medium Risk'

        # ----- Apply visual-similarity decision when available -----------
        v = visual_decisions.get(row_num) or {}
        v_decision = v.get('decision', 'skipped')
        v_cosine = v.get('clip_cosine', -1.0)
        v_phash = v.get('phash_distance', -1)

        # Class overlap is needed for the "identical + Live + same
        # industry → High" branch of Jonathan's rule.
        cls_parts = [int(p) for p in re.split(r'[,\s]+', cleanstr(r[7]))
                     if p.strip().isdigit()]
        class_overlap_n = sum(1 for n in cls_parts if n in target_classes)
        status_l = status_str.lower()

        if v_decision == 'identical':
            # Strong visual signal — apply Jonathan's rule.
            if status_l in ('registered', 'pending'):
                image_risk = 'High Risk' if class_overlap_n > 0 else 'Medium Risk'
            elif status_l == 'ended':
                # Upgrade a Dead cited mark from Negligible to Medium
                # because the logo is identical. Recovers it from the
                # 3f Dead table back into the live 3e Logo table.
                image_risk = 'Medium Risk'
            # else (rare unknown status): leave base
        elif v_decision in ('weak', 'unrelated'):
            # CLIP definitively says different — demote.
            if status_l != 'ended':
                image_risk = 'Low Risk'
            # else: stay Negligible (unchanged)
        # 'similar' or 'skipped': keep the Medium-capped base score.

        # ----- BR-IMG-003: Client Likely override (10 Jun 2026) -----------
        # If the cited mark's OWNER fuzzy-matches the client's company
        # name, this is almost certainly the client's own registered mark
        # appearing in the search results. Override BOTH axes' risk grade
        # to the 'Client Likely' tag so the operator can verify it's
        # really them rather than confusing it with a real third-party
        # threat. The row is still emitted (the user explicitly wants
        # visibility, not exclusion) — it just gets a distinct tag that
        # sorts to the top of the table and is coloured differently.
        client_likely = is_likely_client_owner(client_name, cleanstr(r[8]))
        if client_likely:
            word_risk = 'Client Likely'
            image_risk = 'Client Likely'

        base_record = {
            'office': cleanstr(r[0]),
            'app': cleanstr(r[1]),
            'status': status_str,
            'type': cleanstr(r[3]),
            'mark': cleanstr(r[5]),
            'filing': cleanstr(r[6]),
            'classes': cleanstr(r[7]),
            'owner': cleanstr(r[8]),
            'industry': cleanstr(r[11]),
            'goods': cleanstr(r[12]),
            'image_bytes': images.get(row_num),
            'source_row': row_num,
            # Always expose BOTH scores so downstream code can show them
            # side by side when useful, even in single-axis reports.
            'word_score': word_score, 'word_risk': word_risk,
            'image_score': image_score, 'image_risk': image_risk,
            # Visual similarity fields (BR-IMG-002). 'visual_decision' is
            # the rule input; the numeric fields back it up so the report
            # builder can render a per-row 'why this grade' explanation.
            'visual_decision': v_decision,
            'visual_phash_distance': v_phash,
            'visual_clip_cosine': v_cosine,
            # BR-IMG-003: per-record flag so the report builder can apply
            # the Client-Likely cell colour without re-running the match.
            'client_likely': client_likely,
        }

        def _emit(record_axis: str, score: int, risk: str):
            entry = dict(base_record)
            entry['score'] = score
            entry['risk'] = risk
            entry['threat_type'] = record_axis  # 'Word' or 'Logo'
            scored.append(entry)

        # ----------- Axis-eligibility gates (10 Jun 2026) ------------
        # A record only emits a Word-axis row if its mark text actually
        # matches one of the operator's word_searches rows. Without this
        # gate every figurative mark sharing a class with the client got a
        # Word threat row even when the mark text had zero overlap (e.g.
        # 'B GROUPE BRANDT' scored Low Risk on Word axis just from class
        # overlap, polluting Section 1 with hundreds of false positives).
        #
        # A record only emits a Logo-axis row if its Type contains a
        # figurative element (Figurative / Combined / Stylised). A pure
        # 'Word' mark cannot pose a logo threat, even if it shares a class.
        #
        # When word_searches is empty (legacy callers), we fall back to the
        # pre-fix behaviour and allow Word emission unconditionally so the
        # older single-axis pipelines still work.
        word_eligible = (not word_searches) or mark_matches_any(base_record['mark'], word_searches)
        logo_eligible = is_figurative_type(base_record['type'])

        if report_type == 'word_only':
            if word_eligible:
                _emit('Word', word_score, word_risk)
        elif report_type == 'logo':
            if logo_eligible:
                _emit('Logo', image_score, image_risk)
        else:  # combined
            word_active = word_eligible and (word_risk != 'Negligible')
            image_active = logo_eligible and (image_risk != 'Negligible')
            if word_active and image_active:
                # Both axes apply — emit both so each threat is reviewed
                # independently.
                _emit('Word', word_score, word_risk)
                _emit('Logo', image_score, image_risk)
            elif word_active:
                _emit('Word', word_score, word_risk)
            elif image_active:
                _emit('Logo', image_score, image_risk)
            else:
                # Neither axis active. If at least one axis is *eligible*
                # but Negligible (e.g. status=Ended), keep the record on
                # the higher-scoring eligible axis so the Dead-records
                # section 3f is still populated. If NEITHER axis is
                # eligible (no word match AND not figurative), drop the
                # record — it's a false positive shared-class match.
                if word_eligible and logo_eligible:
                    if image_score > word_score:
                        _emit('Logo', image_score, image_risk)
                    else:
                        _emit('Word', word_score, word_risk)
                elif word_eligible:
                    _emit('Word', word_score, word_risk)
                elif logo_eligible:
                    _emit('Logo', image_score, image_risk)
                # else: drop — no reason to flag this record at all.

    # Sort: live (non-Negligible) first by score desc, then by mark text.
    # Ties between Word and Logo duplicates keep their natural order.
    # Sort: Client Likely first (operator sees their own mark up top to
    # verify), then live rows by score desc, Negligible last.
    def _risk_sort_key(risk: str) -> int:
        if risk == 'Client Likely':
            return 0
        if risk == 'Negligible':
            return 2
        return 1
    scored.sort(key=lambda x: (_risk_sort_key(x['risk']), -x['score'],
                                x['mark'], x.get('threat_type', '')))
    return scored


def process_companies(rows: list, target_sic: str = '45320', root: str = 'STEALTH',
                      word_searches: list[dict] | None = None) -> list[dict]:
    """Process the Companies House sheet, filtering by SIC + word_searches.

    Prefers the operator's `word_searches` list (type-aware matching) when
    supplied; falls back to the legacy `root`-prefix logic otherwise.
    """
    data = [r for r in rows[1:] if r[0] is not None]
    seen = set()
    deduped = []
    for r in data:
        key = cleanstr(r[4])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)
    out = []
    # word_searches no longer filters records (10 Jun 2026, see process_trademarks
    # comment for the full rationale). SIC overlap remains the primary
    # company-side relevance filter.
    use_legacy_root_filter = (not word_searches) and root and root != 'STEALTH'
    for r in deduped:
        mark = cleanstr(r[1])
        sic = cleanstr(r[5]) if len(r) > 5 else ''
        if use_legacy_root_filter and not mark_in_scope_for_root(mark, root):
            continue
        if not has_sic(sic, target_sic):
            continue
        out.append({
            'mark': cleanstr(r[1]),
            'link': cleanstr(r[2]),
            'status': cleanstr(r[3]),
            'number': cleanstr(r[4]),
            'sic': sic,
            'risk': 'Low Risk',
        })
    return out


def process_google(rows: list, images: dict[int, bytes] | None = None) -> list[dict]:
    """Process the Google sheet into scored records.

    `images` is the {row_number: image_bytes} map returned by
    extract_google_image_cells(); when supplied, any data row whose
    column-A cell carries a rich-value (Picture-in-Cell) image gets an
    `image_bytes` field so the report can render the keyword cell as an
    inline image instead of the literal '#VALUE!' Excel exposes.

    Inherits-from-prior keyword fallback: if column A is an Excel error
    value (#VALUE!, #REF!, etc.), the displayed keyword falls back to the
    last non-error value seen in the sheet (typically the parent search
    term).
    """
    images = images or {}
    # Excel error indicators that mean "this cell did not evaluate" — fall
    # back to inheriting the previous valid keyword for display.
    EXCEL_ERRORS = {'#VALUE!', '#REF!', '#N/A', '#NAME?', '#NULL!', '#NUM!', '#DIV/0!'}
    last_valid_keyword = ''
    out = []
    # Rows in the sheet are 1-indexed and row 1 is the header, so the first
    # data row in the spreadsheet is row 2.
    for spreadsheet_row, r in enumerate(rows[1:], start=2):
        if r[0] is None and (len(r) < 2 or r[1] is None):
            continue
        kw_raw = cleanstr(r[0])
        mark = cleanstr(r[1]) if len(r) > 1 else ''
        link = cleanstr(r[2]) if len(r) > 2 else ''
        img_bytes = images.get(spreadsheet_row)
        # Resolve the displayed keyword:
        #   - if there's an image, the keyword is the image (caller will
        #     render image_bytes; we still set keyword to the inherited value
        #     so the report has a textual fallback if the image fails)
        #   - if kw is a known Excel error, inherit from the previous valid
        #   - otherwise use kw as-is and remember it for next time
        if img_bytes:
            keyword = last_valid_keyword or mark or ''
        elif kw_raw in EXCEL_ERRORS:
            keyword = last_valid_keyword or mark or '(keyword unavailable)'
        else:
            keyword = kw_raw
            if keyword:
                last_valid_keyword = keyword

        if 'led' in (keyword + mark).lower():
            risk, score = 'Medium Risk', 44.35
        else:
            risk, score = 'Low Risk', 30.04
        out.append({
            'keyword': keyword,
            'mark_text': mark,
            'link': link,
            'risk': risk,
            'score': score,
            'image_bytes': img_bytes,
        })
    return out


def process_domains(rows: list,
                    domain_searches: list[dict] | None = None) -> list[dict]:
    """Process the Domains sheet.

    When `domain_searches` is supplied, each record is kept only if at
    least one of its URLs (or the mark text) matches at least one operator
    row using type-aware matching (Exact Match / Starts With / Contains /
    Similar To). Records that don't match any row are dropped — this is
    what makes the operator's edits in the form propagate into the report.

    When `domain_searches` is None / empty, every record is kept (legacy
    behaviour). The risk-band heuristic is unchanged.
    """
    data = [r for r in rows[1:] if r[0] is not None]
    out = []
    for r in data:
        mark = cleanstr(r[0])
        urls = [cleanstr(r[i]) for i in range(1, 6) if i < len(r) and r[i]]
        urls = [u for u in urls if u]
        if not urls:
            continue
        if domain_searches:
            # Check the mark text and every URL — keep the record if any of
            # them match any operator row.
            if not domain_matches_any(mark, domain_searches):
                if not any(domain_matches_any(u, domain_searches) for u in urls):
                    continue
        is_led = 'led' in (mark + ''.join(urls)).lower()
        risk, score = ('High Risk', 63) if is_led else ('Medium Risk', 40.76)
        out.append({'mark_text': mark, 'urls': urls, 'risk': risk, 'score': score})
    return out


def process_social(rows: list) -> list[dict]:
    data = [r for r in rows[1:] if r[0] is not None]
    out = []
    for r in data:
        mark = cleanstr(r[0])
        plats = {
            'Facebook': cleanstr(r[1]) if len(r) > 1 else '',
            'Instagram': cleanstr(r[2]) if len(r) > 2 else '',
            'LinkedIn': cleanstr(r[3]) if len(r) > 3 else '',
            'TikTok': cleanstr(r[4]) if len(r) > 4 else '',
            'YouTube': cleanstr(r[5]) if len(r) > 5 else '',
            'X': cleanstr(r[6]) if len(r) > 6 else '',
        }
        if not any(plats.values()):
            continue
        is_led = 'led' in mark.lower()
        risk, score = ('High Risk', 63) if is_led else ('Medium Risk', 40.76)
        out.append({'mark_text': mark, 'platforms': plats, 'risk': risk, 'score': score})
    return out
